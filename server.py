from bottle import route,post,template,response,request,static_file,redirect ,run
import win32api
import openpyxl

@route('/')
def index():
    return template('counter')

@route('/static/<filename>')
def server_static(filename):
    return static_file(filename, root='./static/')

@post('/printcoincounter')
def printcoincounter():
    params = request.forms
    r1=int(params["r1"] or 0)
    r5=int(params["r5"] or 0)
    r10=int(params["r10"] or 0)
    r50=int(params["r50"] or 0)
    r100=int(params["r100"] or 0)
    r500=int(params["r500"] or 0)
    r1000=int(params["r1000"] or 0)
    r2000=int(params["r2000"] or 0)
    r5000=int(params["r5000"] or 0)
    r10000=int(params["r10000"] or 0)
    ridealsum=int(params["ridealsum"] or 0)

    c1=int(params["c1"] or 0)
    c5=int(params["c5"] or 0)
    c10=int(params["c10"] or 0)
    c50=int(params["c50"] or 0)
    c100=int(params["c100"] or 0)
    c500=int(params["c500"] or 0)
    c1000=int(params["c1000"] or 0)
    c2000=int(params["c2000"] or 0)
    c5000=int(params["c5000"] or 0)
    c10000=int(params["c10000"] or 0)
    cidealsum=int(params["cidealsum"] or 0)

    wb = openpyxl.load_workbook("kinsyua5calcnocomment.xlsx")
    sheet = wb.worksheets[0]
    sheet.cell(row=7,column=4,value=r1)
    sheet.cell(row=8,column=4,value=r5)
    sheet.cell(row=9,column=4,value=r10)
    sheet.cell(row=10,column=4,value=r50)
    sheet.cell(row=11,column=4,value=r100)
    sheet.cell(row=12,column=4,value=r500)

    sheet.cell(row=17,column=4,value=r10000)
    sheet.cell(row=18,column=4,value=r5000)
    sheet.cell(row=19,column=4,value=r2000)
    sheet.cell(row=20,column=4,value=r1000)

    sheet.cell(row=26,column=6,value=ridealsum)
##############
    sheet.cell(row=7,column=12,value=c1)
    sheet.cell(row=8,column=12,value=c5)
    sheet.cell(row=9,column=12,value=c10)
    sheet.cell(row=10,column=12,value=c50)
    sheet.cell(row=11,column=12,value=c100)
    sheet.cell(row=12,column=12,value=c500)

    sheet.cell(row=17,column=12,value=c10000)
    sheet.cell(row=18,column=12,value=c5000)
    sheet.cell(row=19,column=12,value=c2000)
    sheet.cell(row=20,column=12,value=c1000)

    sheet.cell(row=26,column=14,value=cidealsum)


    wb.save("result.xlsx")

    win32api.ShellExecute(0,"print","result.xlsx",None,".",0)

    return "作成終了"

run(host='localhost',port=8080,debug=True,reloader=True)
